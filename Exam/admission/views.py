from django.shortcuts import render
from .forms import ExcelUploadForm
from .models import Student
import openpyxl

# installations
# pip install django-import-export
# pip install pandas

# Create your views here.

# def upload_excel(request):
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = form.cleaned_data['excel_file']
#             wb = openpyxl.load_workbook(excel_file)
#             sheet = wb.active
#             for row in sheet.iter_rows(values_only=True):
#                 name, score = row
#                 candidate, created = Candidate.objects.get_or_create(name=name, score=score)
#             return render(request, 'admission/upload_success.html')
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'admission/upload_excel.html', {'form': form})

# def upload_excel(request):
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = form.cleaned_data['excel_file']
#             wb = openpyxl.load_workbook(excel_file)
#             sheet = wb.active
#             for row in sheet.iter_rows(values_only=True):
#                 name, score, program = row
#                 candidate, created = Candidate.objects.get_or_create(name=name, score=score, program=program)
#             return render(request, 'admission/upload_success.html')
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'admission/upload_excel.html', {'form': form})



# def grant_admission(request):
#     candidates = Candidate.objects.all().order_by('-score')
#     for idx, candidate in enumerate(candidates):
#         if idx < 10:
#             candidate.admission_status = True
#             candidate.save()
#     return render(request, 'admission/admission_granted.html', {'candidates': candidates})

# def grant_admission(request):
#     programs = {
#         'Engineering': 80,
#         'Medicine': 85,
#         'Business': 75
#     }
    
#     candidates = Candidate.objects.all().order_by('-score')
#     for candidate in candidates:
#         if candidate.score >= programs.get(candidate.program, 0):
#             candidate.admission_status = True
#             candidate.save()
    
#     return render(request, 'admission/admission_granted.html', {'candidates': candidates})



# excel file upload
# views.py
# import pandas as pd
# from django.shortcuts import render
# from django.http import HttpResponse

# def upload_excel(request):
#     if request.method == 'POST' and request.FILES['excel_file']:
#         excel_file = request.FILES['excel_file']
        
#         if excel_file.name.endswith('.xls') or excel_file.name.endswith('.xlsx'):
#             df = pd.read_excel(excel_file)
            
#             # Extract rows and columns values
#             rows = len(df)
#             columns = len(df.columns)
            
#             # Display the extracted values
#             return render(request, 'excel_data.html', {'rows': rows, 'columns': columns})
#         else:
#             return HttpResponse("Please upload a valid Excel file.")
    
#     return render(request, 'upload_excel.html')




# In this updated implementation:
# The view calculates the total score by summing the scores from
# four subjects extracted from the Excel file.
# It defines a cut-off score for admission (e.g., cut_off_score = 200).
# The view determines the admission status for each candidate based on
# whether their total score is greater than or equal to the cut-off score.
# The extracted data, including the total score and admission status, is displayed in
# an HTML table using the to_html() method.
# Ensure that the column names in the Excel file match the subject names used in the
# calculation logic. Customize the view and templates as needed to suit your specific
# requirements and display the admission status effectively.

# views.py
# import pandas as pd
# from django.shortcuts import render
# from django.http import HttpResponse

# def upload_excel(request):
#     if request.method == 'POST' and request.FILES['excel_file']:
#         excel_file = request.FILES['excel_file']
        
#         if excel_file.name.endswith('.xls') or excel_file.name.endswith('.xlsx'):
#             df = pd.read_excel(excel_file)
            
#             # Calculate total score from four subjects
#             df['Total Score'] = df['Subject1'] + df['Subject2'] + df['Subject3'] + df['Subject4']
            
#             # Define cut-off score for admission
#             cut_off_score = 200
            
#             # Check admission status based on total score
#             df['Admission Status'] = df['Total Score'] >= cut_off_score
            
#             # Display the extracted values and admission status
#             return render(request, 'excel_data.html', {'data': df.to_html()})
#         else:
#             return HttpResponse("Please upload a valid Excel file.")
    
#     return render(request, 'upload_excel.html')



# In this updated implementation:
# Cut-off scores are defined for different programs in the cut_off_scores dictionary.
# The view calculates the total score and admission status for each program based on
# the extracted subject scores and the respective cut-off scores.
# The extracted data, total scores, admission status, and program-specific details
# are displayed in an HTML table using the to_html() method.
# You can create HTML templates (excel_data.html) to render the extracted data and
# admission status. Additionally, you can enhance the user interface using JavaScript
# to highlight the admission status based on the cut-off scores for each program.


# To update the code to admit successful candidates, assign them unique
# matriculation numbers, and post their details to the respective departments
# and programs in Django, you can enhance the view logic and implement a mechanism
# to generate matriculation numbers and handle the data transfer. Here's how you can
# achieve this:

# Update the View to Admit Successful Candidates and Generate Matriculation Numbers:

# Identify successful candidates based on meeting the cut-off scores.
# Generate unique matriculation numbers for admitted candidates.
# Implement Data Transfer to Respective Departments and Programs:

# Define a mechanism to post the details of admitted candidates to the respective
# departments and programs.
# Here's an example implementation:

# views.py
import pandas as pd
import random
from django.shortcuts import render
from django.http import HttpResponse

def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xls') or excel_file.name.endswith('.xlsx'):
            return _extracted_from_upload_excel_6(excel_file, request)
        else:
            return HttpResponse("Please upload a valid Excel file.")

    return render(request, 'upload_excel.html')


# TODO Rename this here and in `upload_excel`
def _extracted_from_upload_excel_6(excel_file, request):
    df = pd.read_excel(excel_file)

    # Define cut-off scores for different programs
    cut_off_scores = {
        'ProgramA': 200,
        'ProgramB': 180,
        'ProgramC': 190
    }

    # Calculate total score and check admission status for each program
    for program, cut_off_score in cut_off_scores.items():
        df[f'{program} Total Score'] = df[f'{program} Subject1'] + df[f'{program} Subject2'] + df[f'{program} Subject3'] + df[f'{program} Subject4']
        df[f'{program} Admission Status'] = df[f'{program} Total Score'] >= cut_off_score

            # Generate matriculation numbers for admitted candidates
    df['Matriculation Number'] = [
        f'MAT{random.randint(1000, 9999)}' for _ in range(len(df))
    ]

    # Filter admitted candidates
    admitted_candidates = df[df['ProgramA Admission Status'] & df['ProgramB Admission Status'] & df['ProgramC Admission Status']]

    # Post details of admitted candidates to respective departments and programs
    for index, candidate in admitted_candidates.iterrows():
        # Post candidate details to respective departments and programs
        # Example: post_to_department(candidate)
        pass

    # Display the extracted values, admission status, and matriculation numbers
    return render(request, 'admission_results.html', {'data': admitted_candidates.to_html()})
