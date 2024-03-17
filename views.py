from django.shortcuts import render
from .forms import ExcelUploadForm
from .models import Candidate
import openpyxl
# Create your views here.

# def upload_excel(request):
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = form.cleaned_data['excel_file']
#             wb = openpyxl.load_workbook(excel_file)
#             sheet = wb.active
#             for row in sheet.iter_rows(values_only=True):
#                 name, score = row
#                 candidate, created = Candidate.objects.get_or_create(name=name, score=score)
#             return render(request, 'admission/upload_success.html')
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'admission/upload_excel.html', {'form': form})

def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                name, score, program = row
                candidate, created = Candidate.objects.get_or_create(name=name, score=score, program=program)
            return render(request, 'admission/upload_success.html')
    else:
        form = ExcelUploadForm()
    return render(request, 'admission/upload_excel.html', {'form': form})



def grant_admission(request):
    candidates = Candidate.objects.all().order_by('-score')
    for idx, candidate in enumerate(candidates):
        if idx < 10:
            candidate.admission_status = True
            candidate.save()
    return render(request, 'admission/admission_granted.html', {'candidates': candidates})

def grant_admission(request):
    programs = {
        'Engineering': 80,
        'Medicine': 85,
        'Business': 75
    }
    
    candidates = Candidate.objects.all().order_by('-score')
    for candidate in candidates:
        if candidate.score >= programs.get(candidate.program, 0):
            candidate.admission_status = True
            candidate.save()
    
    return render(request, 'admission/admission_granted.html', {'candidates': candidates})